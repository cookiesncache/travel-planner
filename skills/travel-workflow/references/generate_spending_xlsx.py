#!/usr/bin/env python3
"""Generate a travel-planner spending spreadsheet (<dest>-spending.xlsx).

The spending tracker's money lives here, in a standalone XLSX with LIVE FORMULAS,
not in a hand-recalculated markdown table. See spending-integration.md for the
schema, the source-of-truth rules, and the recalc pre-flight step.

Mixed-currency model (convert-to-home):
  Each booking keeps its NATIVE amount + currency (the real foreign figure,
  preserved and always visible). A per-row FX rate converts to the trip's HOME
  currency in a formula column, and Total / Remaining / Still-due sum the home
  column — so there is one meaningful trip total without losing the foreign
  amounts. Home-currency rows default to rate 1.0.

Contract (see spending-integration.md):
  * Formulas are written as STRINGS — openpyxl does not compute cached results.
    Any spreadsheet app recalculates on open; recalc-validate (e.g. LibreOffice
    headless into a temp dir) if you need to read the totals back programmatically.
  * Regeneration keys on ID: `## Bookings` drives the identity rows; the native
    Amount, the FX rate, and the Trip budget are user-editable INPUTS and are
    preserved by ID across regen so hand-entered values are never clobbered.
  * Formula ranges and the summary block grow with the booking count — no fixed
    row ceiling.

Usage:
  from generate_spending_xlsx import generate_spending_xlsx
  generate_spending_xlsx(bookings, budget=3000, home_currency="USD",
                         out_path="tokyo-spending.xlsx")

  python generate_spending_xlsx.py bookings.json tokyo-spending.xlsx
  # bookings.json: {"budget": 3000, "home_currency": "USD", "bookings": [
  #   {"id":"tyo-bk1","item":"Tokyo flight","conf":"ABC123","pay":"prepaid",
  #    "currency":"USD","amount":450,"rate":1},
  #   {"id":"tyo-bk2","item":"Hotel Shinjuku","conf":"XYZ789","pay":"due-on-arrival",
  #    "currency":"JPY","amount":90000,"rate":0.0067}, ... ]}

  python generate_spending_xlsx.py --demo
"""
from __future__ import annotations

import json
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

# --- layout (single source of the cell layout) -----------------------------
HEADERS = ["ID", "Item", "Confirmation #", "Payment status",
           "Currency", "Amount", "FX→home", "Amount (home)"]
FIRST_DATA_ROW = 2
COL_ID, COL_ITEM, COL_CONF, COL_PAY, COL_CUR, COL_AMT, COL_RATE, COL_HOME = range(1, 9)

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")    # yellow = editable input
FORMULA_FILL = PatternFill("solid", fgColor="D9D9D9")   # grey = calculated
NATIVE_FMT = '#,##0.00'      # no symbol — the Currency column says which
RATE_FMT = '0.0000'
PAY_STATUSES = ("prepaid", "due-on-arrival", "points")

# home-currency display formats for the home column / budget / summary
HOME_FORMATS = {
    "USD": '"$"#,##0.00', "EUR": '"€"#,##0.00', "GBP": '"£"#,##0.00',
    "JPY": '"¥"#,##0', "AUD": '"A$"#,##0.00', "CAD": '"C$"#,##0.00',
    "CHF": '"CHF "#,##0.00', "NZD": '"NZ$"#,##0.00',
}


def home_fmt(code):
    return HOME_FORMATS.get((code or "USD").upper(),
                            '#,##0.00" %s"' % (code or "").upper())


def _num(x):
    """Coerce to float; blank / None / unparseable -> None (treated as unknown)."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _read_existing_inputs(path):
    """Preserve user-edited native Amount, FX rate, and Trip budget from an
    existing file, keyed by ID — robust to where the summary block sits."""
    amounts, rates, budget = {}, {}, None
    if not os.path.exists(path):
        return amounts, rates, budget
    wb = load_workbook(path)
    ws = wb["Spending"] if "Spending" in wb.sheetnames else wb.active
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        rid = ws.cell(row, COL_ID).value
        label = ws.cell(row, COL_RATE).value   # summary labels sit in the FX→home col
        if rid not in (None, ""):
            amounts[str(rid)] = ws.cell(row, COL_AMT).value
            rates[str(rid)] = ws.cell(row, COL_RATE).value
        elif label == "Trip budget":
            budget = ws.cell(row, COL_HOME).value
    return amounts, rates, budget


def generate_spending_xlsx(bookings, budget=None, home_currency="USD",
                           out_path="spending.xlsx", preserve_existing_inputs=True):
    """Write the spending spreadsheet.

    bookings: iterable of dicts with keys id (required), item, conf, pay,
              currency, amount (native; None/blank => unknown => excluded),
              rate (FX to home; defaults to 1.0 for home-currency rows).
    budget:   trip budget in the home currency, or None.
    home_currency: the currency totals roll up to.
    """
    bookings = list(bookings)
    home_currency = (home_currency or "USD").upper()
    preserved_amounts, preserved_rates, preserved_budget = ({}, {}, None)
    if preserve_existing_inputs:
        preserved_amounts, preserved_rates, preserved_budget = _read_existing_inputs(out_path)
    if budget is None:
        budget = preserved_budget
    budget = _num(budget)

    wb = Workbook()
    ws = wb.active
    ws.title = "Spending"

    ws.append(HEADERS)
    for col in range(1, 9):
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).fill = INPUT_FILL

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(PAY_STATUSES),
                        allow_blank=True)
    ws.add_data_validation(dv)

    row = FIRST_DATA_ROW
    for b in bookings:
        bid = b.get("id")
        if bid in (None, ""):
            raise ValueError("each booking needs a non-empty 'id': %r" % (b,))
        bid = str(bid)
        cur = (b.get("currency") or home_currency).upper()
        amount = _num(b.get("amount"))
        if amount is None and bid in preserved_amounts:
            amount = _num(preserved_amounts[bid])
        rate = _num(b.get("rate"))
        if rate is None and bid in preserved_rates:
            rate = _num(preserved_rates[bid])
        if rate is None and cur == home_currency:
            rate = 1.0

        ws.cell(row, COL_ID, bid)
        ws.cell(row, COL_ITEM, b.get("item"))
        ws.cell(row, COL_CONF, b.get("conf"))
        ws.cell(row, COL_PAY, b.get("pay"))
        ws.cell(row, COL_CUR, cur)
        ws.cell(row, COL_AMT, amount)
        ws.cell(row, COL_AMT).number_format = NATIVE_FMT
        ws.cell(row, COL_RATE, rate)
        ws.cell(row, COL_RATE).number_format = RATE_FMT
        # home amount stays blank (excluded) unless BOTH native amount and rate exist
        h = "=IF(OR({a}{r}=\"\",{g}{r}=\"\"),\"\",{a}{r}*{g}{r})".format(
            a="F", g="G", r=row)
        ws.cell(row, COL_HOME, h)
        ws.cell(row, COL_HOME).number_format = home_fmt(home_currency)
        for col in range(COL_ID, COL_HOME):           # A..G are inputs
            ws.cell(row, col).fill = INPUT_FILL
        ws.cell(row, COL_HOME).fill = FORMULA_FILL     # H is a formula
        row += 1

    dv.add("D%d:D%d" % (FIRST_DATA_ROW, max(row - 1, FIRST_DATA_ROW)))

    data_last = max(row - 1, FIRST_DATA_ROW)           # >= first row even if empty
    total_row = data_last + 2
    budget_row = data_last + 3
    remaining_row = data_last + 4
    due_row = data_last + 5
    hf = home_fmt(home_currency)

    def put_summary(r, label, value, is_formula):
        ws.cell(r, COL_RATE, label).font = Font(bold=True)
        c = ws.cell(r, COL_HOME, value)
        c.number_format = hf
        c.font = Font(bold=True)
        c.fill = FORMULA_FILL if is_formula else INPUT_FILL

    put_summary(total_row, "Total spent",
                "=SUM(H%d:H%d)" % (FIRST_DATA_ROW, data_last), True)
    put_summary(budget_row, "Trip budget", budget, False)
    put_summary(remaining_row, "Remaining",
                "=H%d-H%d" % (budget_row, total_row), True)
    put_summary(due_row, "Still due in person",
                '=SUMIF(D%d:D%d,"due-on-arrival",H%d:H%d)'
                % (FIRST_DATA_ROW, data_last, FIRST_DATA_ROW, data_last), True)

    ws.cell(due_row + 2, COL_ITEM,
            "Yellow = editable input. Grey = calculated — don't type over it. "
            "Amount = native currency; Amount (home) converts via FX→home to %s."
            % home_currency)

    for col, width in zip("ABCDEFGH", (12, 30, 16, 16, 9, 12, 9, 14)):
        ws.column_dimensions[col].width = width

    wb.save(out_path)
    return out_path


def _demo():
    bookings = [
        {"id": "tyo-bk1", "item": "Tokyo flight (JAL 123)", "conf": "ABC123",
         "pay": "prepaid", "currency": "USD", "amount": 450, "rate": 1},
        {"id": "tyo-bk2", "item": "Hotel Shinjuku (3 nights)", "conf": "XYZ789",
         "pay": "due-on-arrival", "currency": "JPY", "amount": 90000, "rate": 0.0067},
        {"id": "tyo-bk3", "item": "Robot Restaurant (points)", "conf": "PTS001",
         "pay": "points", "currency": "USD", "amount": None},
        {"id": "tyo-bk4", "item": "Day tour (price TBC)", "conf": "TBC002",
         "pay": "due-on-arrival", "currency": "EUR", "amount": None},
    ]
    out = generate_spending_xlsx(bookings, budget=3000, home_currency="USD",
                                 out_path="demo-spending.xlsx")
    print("wrote %s" % os.path.abspath(out))
    print("Expected after recalc (home USD): Total 1053.00, Remaining 1947.00, "
          "Still due 603.00; 2 amounts pending (1 points, 1 unpriced).")


def main(argv):
    if len(argv) == 2 and argv[1] == "--demo":
        _demo()
        return 0
    if len(argv) != 3:
        print(__doc__)
        return 2
    with open(argv[1], encoding="utf-8") as fh:
        data = json.load(fh)
    out = generate_spending_xlsx(data.get("bookings", []), data.get("budget"),
                                 data.get("home_currency", "USD"), argv[2])
    print("wrote %s" % os.path.abspath(out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
